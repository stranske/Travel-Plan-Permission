"""Export utilities for accounting integrations."""

from __future__ import annotations

import csv
import io
from collections.abc import Callable, Iterable, Iterator
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from urllib.parse import urlencode, urljoin

from .models import ExpenseReport

ExportRow = dict[str, str]


class ExportService:
    """Generate CSV and Excel exports for expense reports."""

    schema = ["date", "vendor", "amount", "category", "cost_center", "receipt_link"]

    def __init__(
        self,
        *,
        receipt_base_url: str = "https://receipts.example.com",
        receipt_signer: Callable[[str, datetime], str] | None = None,
    ) -> None:
        self.receipt_base_url = receipt_base_url.rstrip("/") + "/"
        self.receipt_signer = receipt_signer

    def _validate_batch(self, reports: Iterable[ExpenseReport]) -> list[ExpenseReport]:
        materialized = list(reports)
        if len(materialized) > 100:
            raise ValueError("Batch export supports up to 100 expense reports")
        return materialized

    def _build_filename(self, ext: str, batch_id: str, now: datetime) -> str:
        return f"expense_export_{now.date().isoformat()}_{batch_id}.{ext}"

    def _default_signed_link(self, receipt_url: str, expires_at: datetime) -> str:
        target = urljoin(self.receipt_base_url, receipt_url.lstrip("/"))
        return f"{target}?{urlencode({'expires_at': expires_at.isoformat()})}"

    def _signed_link(self, receipt_url: str, expires_at: datetime) -> str:
        if self.receipt_signer is not None:
            return self.receipt_signer(receipt_url, expires_at)
        return self._default_signed_link(receipt_url, expires_at)

    def _iter_rows(
        self, reports: list[ExpenseReport], now: datetime
    ) -> Iterator[ExportRow]:
        expires_at = now + timedelta(days=7)
        for report in reports:
            for expense in report.expenses:
                amount = expense.amount.quantize(Decimal("0.01"))
                receipt_link = (
                    self._signed_link(expense.receipt_url, expires_at)
                    if expense.receipt_url
                    else ""
                )
                yield {
                    "date": expense.expense_date.isoformat(),
                    "vendor": expense.vendor or "",
                    "amount": f"{amount:.2f}",
                    "category": expense.category.value,
                    "cost_center": report.cost_center or "",
                    "receipt_link": receipt_link,
                }

    def to_csv(
        self,
        reports: Iterable[ExpenseReport],
        *,
        batch_id: str,
        now: datetime | None = None,
    ) -> tuple[str, str]:
        """Return filename and UTF-8 CSV content."""

        current_time = now or datetime.now(UTC)
        materialized = self._validate_batch(reports)
        rows = self._iter_rows(materialized, current_time)

        output = io.StringIO(newline="")
        writer = csv.DictWriter(output, fieldnames=self.schema)
        writer.writeheader()
        writer.writerows(rows)

        filename = self._build_filename("csv", batch_id, current_time)
        return filename, output.getvalue()

    def to_excel(
        self,
        reports: Iterable[ExpenseReport],
        *,
        batch_id: str,
        now: datetime | None = None,
    ) -> tuple[str, bytes]:
        """Return filename and Excel binary content."""

        from openpyxl import Workbook  # type: ignore[import-untyped]

        current_time = now or datetime.now(UTC)
        materialized = self._validate_batch(reports)
        rows = self._iter_rows(materialized, current_time)

        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"
        ws.append(self.schema)
        for row in rows:
            ws.append(
                [
                    row["date"],
                    row["vendor"],
                    float(row["amount"]),
                    row["category"],
                    row["cost_center"],
                    row["receipt_link"],
                ]
            )
            appended_row = ws.max_row
            receipt_cell = ws.cell(row=appended_row, column=len(self.schema))
            if row["receipt_link"]:
                receipt_cell.hyperlink = row["receipt_link"]
                receipt_cell.style = "Hyperlink"
        amount_column = 3
        currency_format = "$#,##0.00"
        for cell in ws.iter_cols(
            min_col=amount_column, max_col=amount_column, min_row=2
        ):
            for amt_cell in cell:
                amt_cell.number_format = currency_format
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 32

        buffer = io.BytesIO()
        wb.save(buffer)

        filename = self._build_filename("xlsx", batch_id, current_time)
        return filename, buffer.getvalue()
