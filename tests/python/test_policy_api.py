import shutil
import subprocess
from collections.abc import Callable
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pytest

import travel_plan_permission.policy_api as policy_api
from openpyxl import Workbook, load_workbook
from travel_plan_permission import (
    ExpenseCategory,
    PolicyCheckResult,
    Receipt,
    ReconciliationResult,
    TripPlan,
    check_trip_plan,
    list_allowed_vendors,
    reconcile,
)
from travel_plan_permission.mapping import TemplateMapping
from travel_plan_permission.policy import PolicyEngine, PolicyResult, PolicyRule, Severity


class _AlwaysPassRule(PolicyRule):
    rule_id = "always_pass"

    def __init__(self) -> None:
        super().__init__(Severity.BLOCKING)

    def evaluate(self, context) -> PolicyResult:
        return self._result(True, "Always passes for test coverage")

    def message(self) -> str:  # pragma: no cover - test-only rule
        return "Always passes for test coverage"


def test_default_template_path_points_to_existing_file() -> None:
    template_path = policy_api._default_template_path()

    assert template_path.is_file()


def test_split_destination_parses_zip_code() -> None:
    city_state, zip_code = policy_api._split_destination("Austin, TX 78701")

    assert city_state == "Austin, TX"
    assert zip_code == "78701"


def test_plan_field_values_include_derived_fields(
    trip_plan_factory: Callable[..., TripPlan],
) -> None:
    plan = trip_plan_factory(destination="Austin, TX 78701")

    fields = policy_api._plan_field_values(plan)

    assert fields["traveler_name"] == plan.traveler_name
    assert fields["business_purpose"] == plan.purpose
    assert fields["city_state"] == "Austin, TX"
    assert fields["destination_zip"] == "78701"


def test_resolve_field_value_handles_nested_paths() -> None:
    data = {
        "hotel": {"nightly_rate": Decimal("199.99")},
        "comparable_hotels": [{"nightly_rate": Decimal("189.00")}],
    }

    assert (
        policy_api._resolve_field_value(data, "hotel.nightly_rate")
        == Decimal("199.99")
    )
    assert (
        policy_api._resolve_field_value(
            data, "comparable_hotels[0].nightly_rate"
        )
        == Decimal("189.00")
    )
    assert policy_api._resolve_field_value(data, "comparable_hotels[1].nightly_rate") is None


def test_format_helpers_handle_expected_types() -> None:
    assert policy_api._format_date_value(date(2024, 9, 15)) == "2024-09-15"
    assert (
        policy_api._format_date_value(datetime(2024, 9, 15, 8, 0))
        == "2024-09-15"
    )
    assert policy_api._format_date_value("2024-09-15") == "2024-09-15"
    assert policy_api._format_date_value(123) is None

    assert policy_api._format_currency_value(Decimal("10.125")) == Decimal("10.12")
    assert policy_api._format_currency_value(12) == Decimal("12.00")
    assert policy_api._format_currency_value(12.345) == Decimal("12.34")
    assert policy_api._format_currency_value("9.5") == Decimal("9.50")
    assert policy_api._format_currency_value("not-a-number") is None
    assert policy_api._format_currency_value(None) is None


def test_fill_travel_spreadsheet_populates_mapped_fields(
    tmp_path, trip_plan_factory: Callable[..., TripPlan], monkeypatch
) -> None:
    template_path = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.save(template_path)

    plan = trip_plan_factory(
        transportation_mode="air",
        expected_costs={},
        expense_breakdown={ExpenseCategory.CONFERENCE_FEES: Decimal("250.00")},
    )
    mapping = TemplateMapping(
        version="TEST-1",
        cells={
            "traveler_name": "A1",
            "depart_date": "A2",
            "event_registration_cost": "A3",
        },
        dropdowns={"transportation_mode": {"cell": "B1"}},
        checkboxes={
            "expected_costs": {
                "cell": "B2",
                "true_value": "YES",
                "false_value": "NO",
            }
        },
        formulas={"total": {"cell": "C1", "formula": "=A3*1"}},
        metadata={},
    )

    monkeypatch.setattr(policy_api, "load_template_mapping", lambda: mapping)
    monkeypatch.setattr(policy_api, "_default_template_path", lambda *_: template_path)

    output_path = tmp_path / "filled.xlsx"
    policy_api.fill_travel_spreadsheet(plan, output_path)

    filled = load_workbook(output_path)
    sheet = filled.active

    assert sheet["A1"].value == plan.traveler_name
    assert sheet["A2"].value == "2024-09-15"
    assert sheet["A3"].value == 250.0
    assert sheet["A3"].number_format == "$#,##0.00"
    assert sheet["B1"].value == "air"
    assert sheet["B2"].value == "NO"
    assert sheet["C1"].value == "=A3*1"
    filled.close()


def test_check_trip_plan_reports_policy_issues(base_trip_plan: TripPlan) -> None:
    result = check_trip_plan(base_trip_plan)

    assert isinstance(result, PolicyCheckResult)
    assert result.policy_version
    assert result.status == "fail"
    assert any(issue.code == "fare_evidence" for issue in result.issues)
    assert any(issue.code == "hotel_comparison" for issue in result.issues)
    for issue in result.issues:
        assert issue.context["rule_id"] == issue.code


def test_check_trip_plan_passes_with_compliant_engine(
    monkeypatch, base_trip_plan: TripPlan
) -> None:
    engine = PolicyEngine([_AlwaysPassRule()])
    monkeypatch.setattr(
        policy_api.PolicyEngine, "from_file", lambda *args, **kwargs: engine
    )

    result = check_trip_plan(base_trip_plan)

    assert isinstance(result, PolicyCheckResult)
    assert result.policy_version
    assert result.status == "pass"
    assert result.issues == []


def test_check_trip_plan_handles_missing_fields(
    trip_plan_factory: Callable[..., TripPlan],
) -> None:
    plan = trip_plan_factory(expense_breakdown={})

    result = check_trip_plan(plan)

    assert isinstance(result, PolicyCheckResult)
    assert result.status == "fail"
    assert any(issue.code == "fare_evidence" for issue in result.issues)


def test_list_allowed_vendors_returns_registry_matches(
    base_trip_plan: TripPlan,
) -> None:
    vendors = list_allowed_vendors(base_trip_plan)

    assert isinstance(vendors, list)
    assert all(isinstance(vendor, str) for vendor in vendors)
    assert vendors == [
        "Blue Skies Airlines",
        "Citywide Rides",
        "Downtown Suites",
    ]


def test_list_allowed_vendors_filters_destination(
    trip_plan_factory: Callable[..., TripPlan],
) -> None:
    plan = trip_plan_factory(destination="Denver, CO")

    vendors = list_allowed_vendors(plan)

    assert isinstance(vendors, list)
    assert vendors == ["Citywide Rides"]


def test_list_allowed_vendors_filters_by_date(
    trip_plan_factory: Callable[..., TripPlan],
) -> None:
    plan = trip_plan_factory(
        destination="Seattle, WA",
        departure_date=date(2023, 6, 10),
        return_date=date(2023, 6, 12),
    )

    vendors = list_allowed_vendors(plan)

    assert isinstance(vendors, list)
    assert vendors == ["Harbor Hotels"]


def test_reconcile_summarizes_receipts_over_budget(
    base_trip_plan: TripPlan, sample_receipts: list
) -> None:
    result = reconcile(base_trip_plan, sample_receipts)

    assert isinstance(result, ReconciliationResult)
    assert result.status == "over_budget"
    assert result.receipt_count == 2
    assert result.receipts_by_type == {".pdf": 1, ".png": 1}
    assert result.expenses_by_category == {ExpenseCategory.OTHER: Decimal("1200.00")}


def test_reconcile_on_budget_with_matching_receipts(
    trip_plan_factory: Callable[..., TripPlan], sample_receipts: list
) -> None:
    plan = trip_plan_factory(estimated_cost=Decimal("1200.00"))

    result = reconcile(plan, sample_receipts)

    assert isinstance(result, ReconciliationResult)
    assert result.status == "on_budget"
    assert result.variance == Decimal("0")
    assert result.actual_total == Decimal("1200.00")
    assert result.planned_total == Decimal("1200.00")


def test_reconcile_under_budget_with_no_receipts(
    trip_plan_factory: Callable[..., TripPlan],
) -> None:
    plan = trip_plan_factory(estimated_cost=Decimal("800.00"))

    result = reconcile(plan, [])

    assert isinstance(result, ReconciliationResult)
    assert result.status == "under_budget"
    assert result.receipt_count == 0
    assert result.receipts_by_type == {}
    assert result.expenses_by_category == {}


def test_policy_api_documentation_examples_match_models() -> None:
    trip_plan_payload = {
        "trip_id": "TRIP-1001",
        "traveler_name": "Alex Rivera",
        "traveler_role": "Senior Analyst",
        "department": "Finance",
        "destination": "Chicago, IL 60601",
        "origin_city": "Austin, TX",
        "destination_city": "Chicago, IL",
        "departure_date": "2025-06-10",
        "return_date": "2025-06-12",
        "purpose": "Quarterly planning summit",
        "transportation_mode": "air",
        "expected_costs": {"airfare": 420.50, "lodging": 600.00},
        "funding_source": "FIN-OPS",
        "estimated_cost": 1200.50,
        "status": "submitted",
        "expense_breakdown": {"airfare": 420.50, "lodging": 600.00, "meals": 180.00},
        "selected_providers": {"airfare": "Skyway Air", "lodging": "Lakeside Hotel"},
        "validation_results": [],
        "approval_history": [],
        "exception_requests": [],
    }

    plan = TripPlan.model_validate(trip_plan_payload)
    assert plan.trip_id == "TRIP-1001"

    policy_result_payload = {
        "status": "fail",
        "issues": [
            {
                "code": "advance_booking",
                "message": "Flights must be booked 14 days in advance",
                "severity": "warning",
                "context": {"rule_id": "advance_booking", "severity": "advisory"},
            }
        ],
        "policy_version": "d7a6d25a",
    }

    policy_result = PolicyCheckResult.model_validate(policy_result_payload)
    assert policy_result.status == "fail"
    assert policy_result.issues[0].code == "advance_booking"

    reconciliation_payload = {
        "trip_id": "TRIP-1004",
        "report_id": "TRIP-1004-reconciliation",
        "planned_total": 900.00,
        "actual_total": 325.25,
        "variance": -574.75,
        "status": "under_budget",
        "receipt_count": 2,
        "receipts_by_type": {".pdf": 1, ".png": 1},
        "expenses_by_category": {"other": 325.25},
    }

    reconciliation_result = ReconciliationResult.model_validate(reconciliation_payload)
    assert reconciliation_result.receipt_count == 2
    assert reconciliation_result.status == "under_budget"
    assert ExpenseCategory.OTHER in reconciliation_result.expenses_by_category


def test_policy_api_markdown_lint() -> None:
    root = Path(__file__).resolve().parents[2]
    local_bin = root / "node_modules" / ".bin" / "markdownlint-cli2"
    lint_bin = local_bin if local_bin.exists() else shutil.which("markdownlint-cli2")
    if not lint_bin:
        pytest.skip("markdownlint-cli2 is not installed")

    subprocess.run(
        [str(lint_bin), "docs/policy-api.md"],
        check=True,
        cwd=root,
    )
