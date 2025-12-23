"""Tests for accounting export service."""

from __future__ import annotations

from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from urllib.parse import parse_qs, urlparse

import pytest
from openpyxl import load_workbook

from travel_plan_permission import ExportService
from travel_plan_permission.models import (
    ApprovalStatus,
    ExpenseCategory,
    ExpenseItem,
    ExpenseReport,
)


def _sample_report() -> ExpenseReport:
    return ExpenseReport(
        report_id="EXP-100",
        trip_id="TRIP-100",
        traveler_name="Terry Traveler",
        cost_center="ENG",
        approval_status=ApprovalStatus.PENDING,
        expenses=[
            ExpenseItem(
                category=ExpenseCategory.MEALS,
                description="Client dinner",
                vendor="Bistro Cafe",
                amount=Decimal("125.50"),
                expense_date=date(2025, 1, 12),
                receipt_attached=True,
                receipt_url="/receipts/abc123",
            )
        ],
    )


class TestExportService:
    """Accounting export behavior."""

    def test_csv_header_and_column_order(self) -> None:
        """CSV export should be UTF-8 with expected header order."""
        service = ExportService()
        now = datetime(2025, 1, 20, 10, 0, tzinfo=UTC)
        filename, content = service.to_csv([_sample_report()], batch_id="batch-1", now=now)

        assert filename == "expense_export_2025-01-20_batch-1.csv"
        header = content.splitlines()[0]
        assert header == "date,vendor,amount,category,cost_center,receipt_link"
        # Ensure UTF-8 encodes without errors
        content.encode("utf-8")

    def test_receipt_link_expiry_is_7_days(self) -> None:
        """Receipt links should expire in exactly 7 days."""
        service = ExportService()
        now = datetime(2025, 2, 1, 8, 30, tzinfo=UTC)
        _, content = service.to_csv([_sample_report()], batch_id="batch-2", now=now)
        row = content.splitlines()[1].split(",")
        receipt_link = row[-1]

        parsed = urlparse(receipt_link)
        params = parse_qs(parsed.query)
        expires_at = params["expires_at"][0]
        expires_dt = datetime.fromisoformat(expires_at)

        assert expires_dt - now == timedelta(days=7)
        assert parsed.scheme in {"http", "https"}

    def test_excel_currency_and_hyperlink(self) -> None:
        """Excel export should format amount as currency and set clickable hyperlinks."""
        service = ExportService()
        now = datetime(2025, 3, 5, 12, 0, tzinfo=UTC)
        filename, content = service.to_excel([_sample_report()], batch_id="batch-3", now=now)

        assert filename == "expense_export_2025-03-05_batch-3.xlsx"
        workbook = load_workbook(BytesIO(content))
        sheet = workbook.active

        assert [cell.value for cell in sheet[1]] == service.schema
        amount_cell = sheet.cell(row=2, column=3)
        assert amount_cell.number_format == "$#,##0.00"
        receipt_cell = sheet.cell(row=2, column=6)
        assert receipt_cell.hyperlink is not None
        assert receipt_cell.hyperlink.target.startswith("https://")

    def test_batch_size_limit(self) -> None:
        """Batch export should reject more than 100 reports."""
        service = ExportService()
        reports = [_sample_report() for _ in range(101)]

        with pytest.raises(ValueError):
            service.to_csv(reports, batch_id="too-many")

    def test_schema_column_order_matches_acceptance(self) -> None:
        """Schema column order must match documented schema exactly."""
        assert ExportService.schema == [
            "date",
            "vendor",
            "amount",
            "category",
            "cost_center",
            "receipt_link",
        ]
