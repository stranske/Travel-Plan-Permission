import json
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

import travel_plan_permission.policy_api as policy_api
from travel_plan_permission import (
    ExpenseCategory,
    TripPlan,
    fill_travel_spreadsheet,
    render_travel_spreadsheet_bytes,
)
from travel_plan_permission.canonical import (
    CanonicalTripPlan,
    canonical_trip_plan_to_model,
)
from travel_plan_permission.mapping import TemplateMapping


def _plan() -> TripPlan:
    return TripPlan(
        trip_id="TRIP-XL-001",
        traveler_name="Jordan Lee",
        department="FIN-OPS",
        destination="Austin, TX 78701",
        departure_date=date(2024, 9, 15),
        return_date=date(2024, 9, 20),
        purpose="Conference planning",
        estimated_cost=Decimal("1200.00"),
        expense_breakdown={
            ExpenseCategory.CONFERENCE_FEES: Decimal("200"),
            ExpenseCategory.AIRFARE: Decimal("450.50"),
            ExpenseCategory.GROUND_TRANSPORT: Decimal("40"),
        },
    )


def test_travel_spreadsheet_template_loads() -> None:
    template_path = policy_api._default_template_path()
    assert template_path.is_file()

    workbook = load_workbook(template_path)

    assert workbook.sheetnames
    workbook.close()


def test_fill_travel_spreadsheet_writes_mapped_fields(tmp_path) -> None:
    plan = _plan()
    output_path = tmp_path / "filled.xlsx"

    result = fill_travel_spreadsheet(plan, output_path)

    assert result == output_path
    workbook = load_workbook(output_path)
    sheet = workbook["Itinerary Form"]

    assert sheet["C6"].value == plan.traveler_name
    assert sheet["C7"].value == plan.purpose
    assert sheet["G6"].value == 78701
    assert sheet["M7"].value == datetime.combine(plan.departure_date, datetime.min.time())
    assert sheet["M8"].value == datetime.combine(plan.return_date, datetime.min.time())
    assert sheet["M10"].value == 200.0
    assert "$" in sheet["M10"].number_format
    assert "0.00" in sheet["M10"].number_format
    assert sheet["M17"].value == 450.5
    assert "$" in sheet["M17"].number_format
    assert "0.00" in sheet["M17"].number_format
    assert sheet["E20"].value == 450.5
    assert "$" in sheet["E20"].number_format
    assert "0.00" in sheet["E20"].number_format
    workbook.close()


def test_render_travel_spreadsheet_bytes_returns_xlsx_bytes() -> None:
    plan = _plan()

    output_bytes = render_travel_spreadsheet_bytes(plan)

    workbook = load_workbook(BytesIO(output_bytes), read_only=True)
    assert workbook.sheetnames
    workbook.close()


def test_fill_travel_spreadsheet_matches_rendered_bytes(tmp_path) -> None:
    plan = _plan()
    output_path = tmp_path / "filled-match.xlsx"

    output_bytes = render_travel_spreadsheet_bytes(plan)
    fill_travel_spreadsheet(plan, output_path)

    # Compare workbook contents instead of raw bytes (which include timestamps)
    wb_from_bytes = load_workbook(BytesIO(output_bytes), read_only=True)
    wb_from_file = load_workbook(output_path, read_only=True)

    assert wb_from_bytes.sheetnames == wb_from_file.sheetnames
    sheet_bytes = wb_from_bytes["Itinerary Form"]
    sheet_file = wb_from_file["Itinerary Form"]
    for cell_ref in ("C6", "C7", "G6", "M7", "M8", "M10", "M17", "E20"):
        assert sheet_bytes[cell_ref].value == sheet_file[cell_ref].value

    wb_from_bytes.close()
    wb_from_file.close()


def test_fill_travel_spreadsheet_rounds_currency_values(tmp_path) -> None:
    plan = _plan().model_copy(
        update={
            "expense_breakdown": {
                ExpenseCategory.AIRFARE: Decimal("450.567"),
            }
        }
    )
    output_path = tmp_path / "filled-rounded.xlsx"

    fill_travel_spreadsheet(plan, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook.active

    assert sheet["M17"].value == 450.57
    assert "$" in sheet["M17"].number_format
    assert "0.00" in sheet["M17"].number_format
    workbook.close()


def test_fill_travel_spreadsheet_does_not_modify_template(tmp_path) -> None:
    template_path = policy_api._default_template_path()
    template_bytes = template_path.read_bytes()
    plan = _plan()
    output_path = tmp_path / "filled.xlsx"

    fill_travel_spreadsheet(plan, output_path)

    assert output_path.is_file()
    assert template_path.read_bytes() == template_bytes


def test_fill_travel_spreadsheet_uses_mapping_cells(tmp_path, monkeypatch) -> None:
    plan = _plan()
    output_path = tmp_path / "filled-custom.xlsx"
    mapping = TemplateMapping(
        version="ITIN-2025.1",
        cells={
            "traveler_name": "Z99",
            "depart_date": "Z100",
            "event_registration_cost": "Z101",
        },
        dropdowns={},
        checkboxes={},
        formulas={},
        metadata={},
    )

    monkeypatch.setattr(policy_api, "load_template_mapping", lambda: mapping)

    fill_travel_spreadsheet(plan, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook.active

    assert sheet["Z99"].value == plan.traveler_name
    assert sheet["Z100"].value == datetime(2024, 9, 15)
    assert sheet["Z101"].value == 200.0
    assert sheet["Z101"].number_format == "$#,##0.00"
    workbook.close()


@pytest.mark.filterwarnings("ignore:.*PydanticSerializationUnexpectedValue.*")
def test_unfilled_mapping_report_tracks_missing_and_invalid(monkeypatch) -> None:
    plan = TripPlan.model_construct(
        trip_id="TRIP-XL-REPORT",
        traveler_name="Jordan Lee",
        destination="Austin, TX 78701",
        departure_date=123,
        return_date=123,
        purpose="Conference planning",
        estimated_cost=Decimal("1200.00"),
        expense_breakdown={ExpenseCategory.CONFERENCE_FEES: object()},
    )
    mapping = TemplateMapping(
        version="ITIN-2025.1",
        cells={
            "event_registration_cost": "B2",
            "depart_date": "B3",
            "nonexistent_field": "B4",
        },
        dropdowns={},
        checkboxes={},
        formulas={},
        metadata={},
    )
    report = policy_api.UnfilledMappingReport()

    monkeypatch.setattr(policy_api, "load_template_mapping", lambda: mapping)

    policy_api.render_travel_spreadsheet_bytes(plan, report=report)

    observed = {(entry.field, entry.reason) for entry in report.cells}
    assert ("event_registration_cost", "invalid_currency") in observed
    assert ("depart_date", "invalid_date") in observed
    assert ("nonexistent_field", "missing") in observed


def test_fill_travel_spreadsheet_uses_template_metadata(tmp_path, monkeypatch) -> None:
    plan = _plan()
    output_path = tmp_path / "filled-template.xlsx"
    template_path = policy_api._default_template_path()
    template_bytes = template_path.read_bytes()
    observed: dict[str, object] = {}

    def fake_default_template_bytes(template_file: str | None = None):
        observed["template_file"] = template_file
        return template_bytes

    mapping = TemplateMapping(
        version="ITIN-2025.1",
        cells={"traveler_name": "B3"},
        dropdowns={},
        checkboxes={},
        formulas={},
        metadata={"template_file": "custom_template.xlsx"},
    )

    monkeypatch.setattr(policy_api, "load_template_mapping", lambda: mapping)
    monkeypatch.setattr(policy_api, "_default_template_bytes", fake_default_template_bytes)

    fill_travel_spreadsheet(plan, output_path)

    assert observed["template_file"] == "custom_template.xlsx"


def test_fill_travel_spreadsheet_uses_canonical_fields(tmp_path) -> None:
    fixture_path = (
        Path(__file__).resolve().parents[1] / "fixtures" / "sample_trip_plan_minimal.json"
    )
    payload = json.loads(fixture_path.read_text(encoding="utf-8"))
    canonical_plan = CanonicalTripPlan.model_validate(payload)
    trip_plan = canonical_trip_plan_to_model(canonical_plan)
    output_path = tmp_path / "filled-canonical.xlsx"

    fill_travel_spreadsheet(trip_plan, output_path, canonical_plan=canonical_plan)

    workbook = load_workbook(output_path)
    sheet = workbook["Itinerary Form"]

    assert canonical_plan.hotel is not None
    assert sheet["C36"].value == canonical_plan.hotel.name
    assert sheet["C38"].value == canonical_plan.hotel.address
    assert sheet["B40"].value == "☒"
    assert sheet["M75"].value == "☒"
    workbook.close()


def test_fill_travel_spreadsheet_populates_flight_and_hotel_preferences(
    tmp_path,
) -> None:
    fixture_path = Path(__file__).resolve().parents[1] / "fixtures" / "sample_trip_plan_rich.json"
    payload = json.loads(fixture_path.read_text(encoding="utf-8"))
    canonical_plan = CanonicalTripPlan.model_validate(payload)
    trip_plan = canonical_trip_plan_to_model(canonical_plan)
    output_path = tmp_path / "filled-rich.xlsx"

    fill_travel_spreadsheet(trip_plan, output_path, canonical_plan=canonical_plan)

    workbook = load_workbook(output_path)
    sheet = workbook["Itinerary Form"]

    assert sheet["F16"].value == "UA204"
    assert sheet["I16"].value == datetime(2025, 11, 12, 9, 10)
    assert sheet["K16"].value == datetime(2025, 11, 12, 12, 45)
    assert sheet["M17"].value == 612.4
    assert sheet["F17"].value == "UA205"
    assert sheet["I17"].value == datetime(2025, 11, 16, 16, 5)
    assert sheet["K17"].value == datetime(2025, 11, 16, 19, 30)
    assert sheet["C36"].value == "Harborview Suites"
    assert sheet["C38"].value == "88 Mission St"
    assert sheet["H38"].value == "San Francisco, CA"
    assert sheet["K36"].value == 289.9
    assert sheet["B40"].value == "☐"
    assert sheet["K44"].value == "Conference hotel was $60 more per night"
    assert sheet["C45"].value == "Market Square Inn"
    assert sheet["H45"].value == 249.0
    assert sheet["M77"].value == "☒"
    assert sheet["D32"].value == "Need early check-in and airport pickup."
    workbook.close()


def test_fill_travel_spreadsheet_populates_mileage_and_air_comparison(tmp_path) -> None:
    fixture_path = Path(__file__).resolve().parents[1] / "fixtures" / "sample_trip_plan_rich.json"
    payload = json.loads(fixture_path.read_text(encoding="utf-8"))
    payload["parking_estimate"] = "48.00"
    payload["ground_transport_estimate"] = "90.00"
    payload["ground_transport"] = {
        "mileage_planned": True,
        "mileage_miles": "212.4",
        "rideshare_planned": True,
        "rideshare_cost": "90.00",
        "rental_planned": False,
    }
    canonical_plan = CanonicalTripPlan.model_validate(payload)
    trip_plan = canonical_trip_plan_to_model(canonical_plan)
    output_path = tmp_path / "filled-ground.xlsx"

    fill_travel_spreadsheet(trip_plan, output_path, canonical_plan=canonical_plan)

    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["Itinerary Form"]
    assert sheet["C73"].value == 212.4
    assert sheet["C74"].value == 0.725
    assert sheet["C75"].value == "=C73*C74"
    assert sheet["I63"].value == 153.99
    assert sheet["I64"].value == 48.0
    assert sheet["I65"].value == 90.0
    assert sheet["M74"].value == "☒"
    assert sheet["N75"].value == 90.0
    workbook.close()


def test_washington_business_canary_prepares_organizational_workbook(tmp_path) -> None:
    fixture_path = (
        Path(__file__).resolve().parents[1]
        / "fixtures"
        / "washington_dc_business_trip.json"
    )
    canonical_plan = CanonicalTripPlan.model_validate_json(
        fixture_path.read_text(encoding="utf-8")
    )
    trip_plan = canonical_trip_plan_to_model(canonical_plan)
    template_path = policy_api._default_template_path()
    template_bytes = template_path.read_bytes()
    output_path = tmp_path / "washington-dc-business-trip.xlsx"

    fill_travel_spreadsheet(trip_plan, output_path, canonical_plan=canonical_plan)

    assert template_path.read_bytes() == template_bytes
    workbook = load_workbook(output_path, read_only=True, data_only=False)
    sheet = workbook["Itinerary Form"]
    expected_values = {
        "C6": "Taylor Morgan",
        "G6": 20001,
        "M6": "10/15/2026 - 10/16/2026",
        "C7": "Client strategy meetings and implementation workshop",
        "M7": datetime(2026, 10, 14),
        "M8": datetime(2026, 10, 16),
        "D14": "St. Louis (STL)",
        "J14": "Washington National (DCA)",
        "F16": "AA 445",
        "F17": "AA 2117",
        "M17": 418.6,
        "E20": 389.1,
        "K22": 36.0,
        "B27": "☒",
        "C36": "Capital Center Hotel (synthetic)",
        "K36": 289.0,
        "B40": "☐",
        "B41": "☒",
        "C45": "District Square Inn (synthetic)",
        "C46": "Penn Quarter Lodge (synthetic)",
        "H52": 2,
        "H53": 2,
        "H54": 2,
        "M75": "☒",
        "N75": 96.0,
        "B107": "☒",
    }
    for cell_ref, expected in expected_values.items():
        assert sheet[cell_ref].value == expected, cell_ref
    assert sheet["H6"].value == (
        '=IFERROR(VLOOKUP(G6,Locations!A:F,4,FALSE)&", "'
        '&VLOOKUP(G6,Locations!A:F,6,FALSE),"City / State")'
    )
    assert sheet["O22"].value == "=SUM(M17,C23,F23,K23)"
    assert sheet["O39"].value == "=K36*O36"
    assert sheet["G103"].value == "=SUM(G97:G102)"
    assert workbook.calculation.fullCalcOnLoad is True
    assert workbook.calculation.forceFullCalc is True
    workbook.close()
