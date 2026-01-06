from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import travel_plan_permission.policy_api as policy_api
from travel_plan_permission import ExpenseCategory, Receipt, TripPlan
from travel_plan_permission.mapping import TemplateMapping
from travel_plan_permission.policy import PolicyEngine, PolicyResult, Severity
from travel_plan_permission.policy_versioning import PolicyVersion


def _blank_template_bytes(initial_cells: dict[str, object] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    for cell, value in (initial_cells or {}).items():
        ws[cell] = value
    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()


@pytest.fixture()
def base_plan() -> TripPlan:
    return TripPlan(
        trip_id="TRIP-HELP-001",
        traveler_name="Taylor Morgan",
        department="FIN",
        destination="Austin, TX 78701",
        departure_date=date(2024, 9, 15),
        return_date=date(2024, 9, 20),
        purpose="Planning session",
        transportation_mode="air",
        estimated_cost=Decimal("1500.00"),
        expense_breakdown={ExpenseCategory.CONFERENCE_FEES: Decimal("0")},
    )


def test_default_template_path_reports_package_resource_only(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    template_bytes = b"template-bytes"

    class FakeResource:
        def joinpath(self, *_parts: str) -> FakeResource:
            return self

        def is_file(self) -> bool:
            return True

        def read_bytes(self) -> bytes:
            return template_bytes

    monkeypatch.setattr(Path, "exists", lambda _self: False)
    monkeypatch.setattr(policy_api.resources, "files", lambda _name: FakeResource())

    template_path = policy_api._default_template_path("missing-template.xlsx")
    assert template_path.name == "missing-template.xlsx"
    assert template_path.read_bytes() == template_bytes


def test_default_template_path_uses_cached_resource(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    template_name = "cached-template.xlsx"
    cached_path = tmp_path / template_name
    cached_path.write_bytes(b"cached")

    class FakeResource:
        def joinpath(self, *_parts: str) -> FakeResource:
            return self

        def is_file(self) -> bool:
            return True

        def read_bytes(self) -> bytes:
            raise AssertionError("Expected cached path to be used.")

    original_cache = dict(policy_api._RESOURCE_TEMPLATE_CACHE)
    policy_api._RESOURCE_TEMPLATE_CACHE.clear()
    policy_api._RESOURCE_TEMPLATE_CACHE[template_name] = cached_path

    def fake_exists(path: Path) -> bool:
        return path == cached_path

    monkeypatch.setattr(Path, "exists", fake_exists)
    monkeypatch.setattr(policy_api.resources, "files", lambda _name: FakeResource())

    try:
        template_path = policy_api._default_template_path(template_name)
        assert template_path == cached_path
    finally:
        policy_api._RESOURCE_TEMPLATE_CACHE.clear()
        policy_api._RESOURCE_TEMPLATE_CACHE.update(original_cache)


def test_default_template_bytes_reads_package_resource(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    template_bytes = b"template-bytes"

    class FakeResource:
        def joinpath(self, *_parts: str) -> FakeResource:
            return self

        def is_file(self) -> bool:
            return True

        def read_bytes(self) -> bytes:
            return template_bytes

    monkeypatch.setattr(Path, "exists", lambda _self: False)
    monkeypatch.setattr(policy_api.resources, "files", lambda _name: FakeResource())

    assert policy_api._default_template_bytes("missing-template.xlsx") == template_bytes


def test_default_template_bytes_missing_package_module(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(Path, "exists", lambda _self: False)

    def _raise_module_error(_name: str):
        raise ModuleNotFoundError

    monkeypatch.setattr(policy_api.resources, "files", _raise_module_error)

    with pytest.raises(FileNotFoundError, match="Unable to locate templates"):
        policy_api._default_template_bytes("missing-template.xlsx")


def test_split_destination_returns_original_when_pattern_misses(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(policy_api, "_ZIP_PATTERN", re.compile(r"^$"))

    city_state, zip_code = policy_api._split_destination("Nowhere Land 99999")

    assert city_state == "Nowhere Land 99999"
    assert zip_code is None


def test_resolve_field_value_handles_nested_variations() -> None:
    assert policy_api._resolve_field_value({"a": 1}, "a") == 1
    assert policy_api._resolve_field_value({"a": 1}, "bad[") is None
    assert policy_api._resolve_field_value({"a": "value"}, "a.b") is None
    assert policy_api._resolve_field_value({"a": {"b": "value"}}, "a.b[0]") is None
    assert policy_api._resolve_field_value({"a": {"b": [1]}}, "a.b[2]") is None
    assert policy_api._resolve_field_value({"a": {"b": [{"c": 3}]}}, "a.b[0].c") == 3


def test_format_date_value_variants() -> None:
    assert policy_api._format_date_value(datetime(2024, 1, 2, 9, 30)) == "2024-01-02"
    assert policy_api._format_date_value(date(2024, 2, 3)) == "2024-02-03"
    assert policy_api._format_date_value("2024-03-04") == "2024-03-04"
    assert policy_api._format_date_value(123) is None


def test_format_currency_value_variants() -> None:
    assert policy_api._format_currency_value(None) is None
    assert policy_api._format_currency_value(10) == Decimal("10.00")
    assert policy_api._format_currency_value(12.5) == Decimal("12.50")
    assert policy_api._format_currency_value("oops") is None
    assert policy_api._format_currency_value(object()) is None


def test_issue_severity_reports_info() -> None:
    result = PolicyResult(
        rule_id="info_rule",
        severity=Severity.INFO,
        passed=True,
        message="Informational",
    )

    assert policy_api._issue_severity(result) == "info"


def test_fill_travel_spreadsheet_sets_dropdown_checkbox_and_formula(
    base_plan: TripPlan, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    mapping = TemplateMapping(
        version="ITIN-2025.1",
        cells={"traveler_name": "A1"},
        dropdowns={"transportation_mode": {"cell": "B1"}},
        checkboxes={"department": {"cell": "C1", "true_value": "Y", "false_value": "N"}},
        formulas={"total": {"cell": "D1", "formula": "=1+1"}},
        metadata={},
    )
    template_bytes = _blank_template_bytes()

    monkeypatch.setattr(policy_api, "load_template_mapping", lambda: mapping)
    monkeypatch.setattr(
        policy_api,
        "_default_template_bytes",
        lambda *_args, **_kwargs: template_bytes,
    )

    output_path = tmp_path / "dropdowns.xlsx"
    policy_api.fill_travel_spreadsheet(base_plan, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook.active
    assert sheet["A1"].value == base_plan.traveler_name
    assert sheet["B1"].value == base_plan.transportation_mode
    assert sheet["C1"].value == "Y"
    assert sheet["D1"].value == "=1+1"
    workbook.close()


def test_fill_travel_spreadsheet_skips_invalid_currency_and_date(
    base_plan: TripPlan, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    mapping = TemplateMapping(
        version="ITIN-2025.1",
        cells={"event_registration_cost": "B2", "depart_date": "C2"},
        dropdowns={},
        checkboxes={},
        formulas={},
        metadata={},
    )
    template_bytes = _blank_template_bytes({"B2": "keep", "C2": "keep"})

    monkeypatch.setattr(policy_api, "load_template_mapping", lambda: mapping)
    monkeypatch.setattr(
        policy_api,
        "_default_template_bytes",
        lambda *_args, **_kwargs: template_bytes,
    )
    monkeypatch.setattr(
        policy_api,
        "_plan_field_values",
        lambda _plan, **_kwargs: {
            "event_registration_cost": "not-a-number",
            "depart_date": object(),
        },
    )

    output_path = tmp_path / "skips.xlsx"
    policy_api.fill_travel_spreadsheet(base_plan, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook.active
    assert sheet["B2"].value == "keep"
    assert sheet["C2"].value == "keep"
    workbook.close()


def test_plan_field_values_includes_cost_center_and_destination(
    base_plan: TripPlan,
) -> None:
    plan = base_plan.model_copy(
        update={
            "department": None,
            "funding_source": "OPS-42",
            "destination": "Denver, CO 80202",
            "expense_breakdown": {
                ExpenseCategory.CONFERENCE_FEES: Decimal("50"),
                ExpenseCategory.AIRFARE: Decimal("275.50"),
                ExpenseCategory.GROUND_TRANSPORT: Decimal("18.75"),
            },
        }
    )

    fields = policy_api._plan_field_values(plan)

    assert fields["traveler_name"] == plan.traveler_name
    assert fields["business_purpose"] == plan.purpose
    assert fields["cost_center"] == "OPS-42"
    assert fields["city_state"] == "Denver, CO"
    assert fields["destination_zip"] == "80202"
    assert fields["depart_date"] == plan.departure_date
    assert fields["return_date"] == plan.return_date
    assert fields["event_registration_cost"] == Decimal("50")
    assert fields["flight_pref_outbound.roundtrip_cost"] == Decimal("275.50")
    assert fields["lowest_cost_roundtrip"] == Decimal("275.50")
    assert fields["parking_estimate"] == Decimal("18.75")


def test_context_from_plan_maps_costs(base_plan: TripPlan) -> None:
    plan = base_plan.model_copy(
        update={
            "expense_breakdown": {
                ExpenseCategory.GROUND_TRANSPORT: Decimal("25"),
                ExpenseCategory.AIRFARE: Decimal("120"),
            }
        }
    )

    context = policy_api._context_from_plan(plan)

    assert context.departure_date == plan.departure_date
    assert context.return_date == plan.return_date
    assert context.driving_cost == Decimal("25")
    assert context.flight_cost == Decimal("120")


def test_policy_version_hash_matches_engine_rules() -> None:
    engine = PolicyEngine([])

    version_hash = policy_api._policy_version(engine)

    expected = PolicyVersion.from_config(None, {"rules": []}).config_hash
    assert version_hash == expected


def test_issue_from_result_formats_context() -> None:
    result = PolicyResult(
        rule_id="blocking_rule",
        severity=Severity.BLOCKING,
        passed=False,
        message="Stop",
    )

    issue = policy_api._issue_from_result(result)

    assert issue.code == "blocking_rule"
    assert issue.message == "Stop"
    assert issue.severity == "error"
    assert issue.context == {"rule_id": "blocking_rule", "severity": "blocking"}


def test_expense_from_receipt_tracks_third_party_payment() -> None:
    receipt = Receipt(
        total=Decimal("45.50"),
        date=date(2024, 10, 2),
        vendor="Cafe Luna",
        file_reference="receipt-005.pdf",
        file_size_bytes=100,
        paid_by_third_party=True,
    )

    expense = policy_api._expense_from_receipt(receipt)

    assert expense.vendor == "Cafe Luna"
    assert expense.receipt_attached is True
    assert expense.receipt_url == "receipt-005.pdf"
    assert expense.third_party_paid_explanation is not None
    assert expense.receipt_references == [receipt]


def test_build_expense_report_aggregates_receipts(base_plan: TripPlan) -> None:
    receipts = [
        Receipt(
            total=Decimal("75.00"),
            date=date(2024, 10, 3),
            vendor="Hotel Luna",
            file_reference="receipt-006.png",
            file_size_bytes=1000,
        )
    ]

    report = policy_api._build_expense_report(base_plan, receipts)

    assert report.report_id == f"{base_plan.trip_id}-reconciliation"
    assert report.trip_id == base_plan.trip_id
    assert report.expenses[0].description == "Receipt from Hotel Luna"
